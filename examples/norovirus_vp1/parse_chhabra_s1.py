"""Extract GenBank accessions from Chhabra 2019 Supplementary Table S1.

The Chhabra et al. 2019 JGV paper (PMC7011714) distributes Table S1 as an xlsx
that is not in PMC's Open Access bulk set, so it must be downloaded manually
from either:

  https://www.microbiologyresearch.org/content/journal/jgv/10.1099/jgv.0.001318
  https://pmc.ncbi.nlm.nih.gov/articles/instance/7011714/bin/jgv-100-1393-s001.xlsx

Save the downloaded file as `chhabra_s1.xlsx` next to this script, then run:

  python parse_chhabra_s1.py --xlsx chhabra_s1.xlsx --out chhabra_2019_vp1_accessions.txt
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("openpyxl is required: pip install openpyxl") from exc

ACC_RE = re.compile(r"\b([A-Z]{1,3}\d{5,8})(?:\.\d+)?\b")


def main() -> int:
    ap = argparse.ArgumentParser(description="Extract accessions from Chhabra 2019 S1.")
    ap.add_argument("--xlsx", type=Path, required=True)
    ap.add_argument("--out", type=Path, required=True)
    args = ap.parse_args()

    wb = load_workbook(args.xlsx, data_only=True)
    accessions: set[str] = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, str):
                    for m in ACC_RE.findall(cell):
                        accessions.add(m)

    with args.out.open("w") as fh:
        for acc in sorted(accessions):
            fh.write(acc + "\n")
    print(f"wrote {len(accessions)} unique accessions to {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
